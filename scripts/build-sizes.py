"""
build-sizes.py — 사이즈 옵션 + 가격(override/증감액) 데이터 빌드.
------------------------------------------------------------------
product_list.xlsx(products 시트)에서:
  1) sizes 컬럼 → lib/catalog/sizes.json
     - 형식: "S, M +2000, L +6000" / "S -50000, M -22000, L" 처럼
       사이즈명 뒤에 +금액·-금액(없으면 0). → [{"name": "S", "delta": 0}, ...]
     - 사이즈명에 괄호·슬래시·단위가 있어도 보존(예: "L/XL", "M(36~51cm)", "1.5in").
  2) price 컬럼 → lib/catalog/prices.json
     - 입력된 제품만 기본가 override({folder: 숫자}). 빈칸은 기존 raw.json 가격 유지.
실행: python scripts/build-sizes.py  (Daengdabang_Shop 루트에서)
"""
import openpyxl, json, re

XLSX = "data/product_list.xlsx"
OUT_SIZES = "lib/catalog/sizes.json"
OUT_PRICES = "lib/catalog/prices.json"

# 사이즈 1개 파싱 — 끝에 " +1234" / "-1234" 가 있으면 증감액, 없으면 0.
SIZE_RE = re.compile(r"^(.+?)\s*([+-]\d+)$")


def parse_size(token: str):
    t = token.strip()
    m = SIZE_RE.match(t)
    if m:
        return {"name": m.group(1).strip(), "delta": int(m.group(2))}
    return {"name": t, "delta": 0}


wb = openpyxl.load_workbook(XLSX)
ws = wb["products"] if "products" in wb.sheetnames else wb.active
H = [c.value for c in ws[1]]
if "sizes" not in H:
    raise SystemExit("[ERROR] 'sizes' 컬럼이 없습니다.")
fi = H.index("folder_name")
si = H.index("sizes")
pi = H.index("price") if "price" in H else None

sizes_result, prices_result = {}, {}
for row in ws.iter_rows(min_row=2, values_only=True):
    folder = row[fi]
    if not folder:
        continue
    raw = row[si]
    if raw and str(raw).strip():
        parsed = [parse_size(s) for s in str(raw).split(",") if s.strip()]
        if parsed:
            sizes_result[folder] = parsed
    if pi is not None:
        price = row[pi]
        if price not in (None, "") and str(price).strip():
            try:
                prices_result[folder] = int(float(str(price).replace(",", "").strip()))
            except ValueError:
                print(f"  [WARN] {folder} price 파싱 실패: {price!r}")

with open(OUT_SIZES, "w", encoding="utf-8") as f:
    json.dump(sizes_result, f, ensure_ascii=False, indent=1)
with open(OUT_PRICES, "w", encoding="utf-8") as f:
    json.dump(prices_result, f, ensure_ascii=False, indent=1)

n_delta = sum(1 for v in sizes_result.values() for s in v if s["delta"] != 0)
print(f"[OK] sizes.json {len(sizes_result)}개 (증감 적용 항목 {n_delta}) / prices.json {len(prices_result)}개")
