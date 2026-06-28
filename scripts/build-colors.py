"""
build-colors.py — 색상 변형 데이터 일괄 빌드.
------------------------------------------------------------------
product_list.xlsx 의 sub_image 컬럼을 읽어 색상 데이터를 만든다.
  1) 각 제품 colors/*.png → *.webp 변환 (색상별 메인 이미지)
  2) 각 제품 colors/chips.png 를 색상 수만큼 분할 → colors/chips/*.webp (색상 칩)
     - 배경(흰색/투명) 제거 후 행→열 투영으로 칩 위치 검출
     - 윗줄 좌→우, 아랫줄 좌→우 순서가 sub_image 순서와 일치 → 칩[i] = 색상[i]
  3) lib/catalog/colors.json 생성 ({folder: [{file, name, chip}]})
칩 검출 수 ≠ 색상 수인 제품은 MISMATCH 로 리포트(개별 점검 필요).
"""
import openpyxl, os, json
from PIL import Image
import numpy as np

BASE = "public/images/products/catalog"
XLSX = "data/product_list.xlsx"
OUT = "lib/catalog/colors.json"


def runs(mask, min_len=10):
    """1D bool 마스크에서 연속 True 구간 [(start,end)] (end 포함). min_len 미만은 노이즈로 버림."""
    res = []
    s = None
    for i, v in enumerate(mask):
        if v and s is None:
            s = i
        elif not v and s is not None:
            if i - s >= min_len:
                res.append((s, i - 1))
            s = None
    if s is not None and len(mask) - s >= min_len:
        res.append((s, len(mask) - 1))
    return res


def split_chips(path):
    """chips.png → 칩 bbox 목록(윗줄 좌→우, 아랫줄 좌→우 순)."""
    img = Image.open(path).convert("RGBA")
    a = np.array(img)
    rgb = a[:, :, :3].astype(int)
    al = a[:, :, 3]
    fg = (rgb.min(axis=2) <= 238) & (al > 30)  # 전경 = 비흰색 & 불투명
    boxes = []
    for r0, r1 in runs(fg.any(axis=1)):  # 행 투영 → 줄
        band = fg[r0:r1 + 1, :]
        for c0, c1 in runs(band.any(axis=0)):  # 각 줄 열 투영 → 칩
            sub = fg[r0:r1 + 1, c0:c1 + 1]
            ys = np.where(sub.any(axis=1))[0]
            xs = np.where(sub.any(axis=0))[0]
            boxes.append((r0 + ys[0], r0 + ys[-1], c0 + xs[0], c0 + xs[-1]))
    return img, boxes


def crop_square(img, box, size=96):
    """칩 bbox 를 정사각형으로 가운데 crop 후 리사이즈(원형 마스크는 CSS 가 처리)."""
    r0, r1, c0, c1 = box
    s = max(r1 - r0 + 1, c1 - c0 + 1)
    cy, cx, h = (r0 + r1) // 2, (c0 + c1) // 2, s // 2
    return img.crop((cx - h, cy - h, cx - h + s, cy - h + s)).resize((size, size), Image.LANCZOS)


wb = openpyxl.load_workbook(XLSX)
ws = wb.active
H = [c.value for c in ws[1]]
si, fi, nidx = H.index("sub_image"), H.index("folder_name"), H.index("no")

result = {}
mism = []
for row in ws.iter_rows(min_row=2, values_only=True):
    sub = row[si]
    if not (sub and str(sub).strip()):
        continue
    folder = row[fi]
    cdir = os.path.join(BASE, folder, "colors")
    parts = [p.strip() for p in str(sub).split(",") if p.strip()]
    items = []  # (png 파일명, 한글명)
    for p in parts:
        fn = p.split("(")[0].strip()
        nm = p[p.find("(") + 1:p.rfind(")")].strip() if "(" in p else ""
        items.append((fn, nm))

    # 1) png → webp (색상별 메인 이미지)
    chipdir = os.path.join(cdir, "chips")
    os.makedirs(chipdir, exist_ok=True)
    entries = []
    for fn, nm in items:
        wn = os.path.splitext(fn)[0] + ".webp"
        Image.open(os.path.join(cdir, fn)).convert("RGB").save(os.path.join(cdir, wn), "WEBP", quality=88)
        entries.append({"file": wn, "name": nm, "chip": f"chips/{wn}"})

    # 2) chips.png 분할
    img, boxes = split_chips(os.path.join(cdir, "chips.png"))
    if len(boxes) == len(items):
        for (fn, nm), box in zip(items, boxes):
            wn = os.path.splitext(fn)[0] + ".webp"
            crop_square(img, box).save(os.path.join(chipdir, wn), "WEBP", quality=90)
    else:
        mism.append(f"no{row[nidx]} {folder}: colors={len(items)} chips_detected={len(boxes)}")

    result[folder] = entries

with open(OUT, "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=1)

print("products:", len(result))
print("MISMATCHES:", len(mism))
for m in mism:
    print("  ", m)
