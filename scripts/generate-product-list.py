"""
scripts/generate-product-list.py
---------------------------------
Daengdabang_Shop/data/product_list.xlsx 를 생성한다.

데이터 소스:
- lib/catalog.json: 333개 카탈로그 (정정 + 재번호 1~333 완료)
- scripts/folder_list.json: Excel 폴더목록 (folder_name 매칭용)
- 원본 Excel (C:\\Users\\lee\\Downloads\\pet_products_brand_purpose_season_detail.xlsx): 고양이 12개 추출용

시트 구성:
- "products" 시트: 333개 강아지 상품 (no = 1~333, 재번호됨)
- "cats" 시트:     12개 고양이 상품 (원본 Excel no 보존, 별도 분류)

컬럼 (products 시트):
  no | product_name | folder_name | main_image | detail_path | detail_url

컬럼 (cats 시트):
  original_no | brand | product_name | folder_name | note

시트 (guide 시트):
  전체 폴더 구조 + 명명 규칙 정리 (작업자 자급자족 매뉴얼)

규칙:
- main_image  = {folder_name}.png
- detail_path = details/1.png, 2.png, ...
- detail_url  = https://www.daengdabang.com/product/{folder_name}
- 갤러리(2.png, 3.png), 사이즈(size.png), 영상(video.mp4)은 자동 감지 (가이드 시트 참고)
- 고양이 시트는 비활성 상품 — 향후 활성화 시 강아지 시트로 옮김
"""
import json
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parent.parent
CATALOG_PATH = ROOT / "lib" / "catalog" / "raw.json"
FOLDER_LIST_PATH = ROOT / "scripts" / "folder_list.json"
ORIGINAL_EXCEL = Path(r"C:\Users\lee\Downloads\pet_products_brand_purpose_season_detail.xlsx")
DATA_DIR = ROOT / "data"
OUT_PATH = DATA_DIR / "product_list.xlsx"

# no=5 신규 부여 (Excel 에 없는 유일한 catalog 항목)
# 다른 릿지라인 시리즈 패턴 따름
NEW_FOLDERS = {
    5: "rw_ridgeline_leash_26",
}

BASE_URL = "https://www.daengdabang.com"

# 고양이 상품 12개 (원본 Excel 의 No)
CAT_NUMBERS = {99, 129, 130, 140, 179, 180, 181, 185, 203, 206, 209, 320}


def norm(s: str) -> str:
    s = (s or "").lower().strip()
    s = re.sub(r"\.\.\.$", "", s)
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[()（）\[\]·.,\-_·•&/!?]", "", s)
    s = s.replace("리프웨어", "러프웨어")
    s = s.replace("팔리세이드", "팰리세이드")
    s = s.replace("핏드림하우스", "펫드림하우스")
    s = s.replace("탐라이프", "탑라이프")
    s = s.replace("퍼그너티", "페그너티")
    s = s.replace("프런트", "프론트")
    s = s.replace("크랙그", "크래그")
    return s


def get_borders():
    return Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )


def write_header(ws, headers, fill_color="4B5563"):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=fill_color)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = get_borders()
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border


def build_products_sheet(ws, catalog, excel_by_norm):
    """products 시트: 깔끔한 표 (헤더 + 333행). 안내는 별도 guide 시트."""
    headers = ["no", "product_name", "folder_name", "main_image", "detail_path", "detail_url"]
    write_header(ws, headers)
    border = get_borders()

    matched_count = 0
    no5_assigned = False

    for c in sorted(catalog, key=lambda x: x["no"]):
        no = c["no"]
        name = c["name"]

        if no == 5:
            folder = NEW_FOLDERS[5]
            no5_assigned = True
        else:
            folder = excel_by_norm.get(norm(name), "")

        if folder:
            matched_count += 1
            main_image = f"{folder}.png"
            detail_path = "details/1.png, 2.png, 3.png, ..."
            detail_url = f"{BASE_URL}/product/{folder}"
        else:
            main_image = ""
            detail_path = ""
            detail_url = f"{BASE_URL}/product/p_{no}"

        row_idx = ws.max_row + 1
        for col_idx, value in enumerate(
            [no, name, folder, main_image, detail_path, detail_url],
            start=1,
        ):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    widths = {1: 6, 2: 60, 3: 32, 4: 28, 5: 30, 6: 60}
    for col_idx, w in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24

    return matched_count


def build_guide_sheet(ws):
    """명명 규칙 + 폴더 구조 안내 시트"""
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="4F46E5")
    section_font = Font(bold=True, size=11, color="111827")
    section_fill = PatternFill("solid", fgColor="EEF2FF")
    body_font = Font(size=10, color="374151")
    code_font = Font(size=10, color="0F172A", name="Consolas")
    border = get_borders()

    # 타이틀
    ws.merge_cells("A1:E1")
    cell = ws.cell(row=1, column=1, value="product_list.xlsx 작업 가이드")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    rows = [
        ("",),
        ("● 작업 위치 (모든 협업자 공통)",),
        ("",),
        ("  각 제품의 자산은 git 저장소 안의 아래 폴더에 보관합니다:",),
        ("  Daengdabang_Shop/public/images/products/catalog/{folder_name}/",),
        ("",),
        ("  folder_name 은 products 시트의 folder_name 컬럼에서 확인.",),
        ("  git clone 으로 받으면 모든 자산이 함께 따라옵니다 — 별도 폴더 공유 불필요.",),
        ("",),
        ("● 폴더 구조 (예시: rw_notarock)",),
        ("",),
        ("  Daengdabang_Shop/public/images/products/catalog/rw_notarock/",),
        ("  ├── rw_notarock.png    ← 메인 이미지 (필수, folder_name + .png)",),
        ("  ├── 2.png              ← 갤러리 2번째 (옵션, 정사각형 권장)",),
        ("  ├── 3.png              ← 갤러리 3번째",),
        ("  ├── 4.png, 5.png, ...  ← 추가 갤러리 (필요한 만큼)",),
        ("  ├── size.png           ← 사이즈 차트 (옵션)",),
        ("  ├── video.mp4          ← 영상 (옵션, mp4 + h.264)",),
        ("  └── details/           ← 상세 페이지 본문 (긴 세로 이미지들)",),
        ("      ├── 1.png          ← 상세 1번째 (헤더/배너)",),
        ("      ├── 2.png          ← 상세 2번째 (제품 소개)",),
        ("      ├── 3.png          ← 상세 3번째 (특징)",),
        ("      └── ...            ← 필요한 만큼 (사이트가 1.png 부터 순서대로 표시)",),
        ("",),
        ("● 파일 명명 규칙",),
        ("",),
        ("  메인 이미지: folder_name 과 동일 + .png",),
        ("              예) folder_name = 'rw_notarock' → rw_notarock.png",),
        ("  갤러리:     2.png 부터 시작 (메인이 1번째이므로)",),
        ("  상세 본문:  details/ 폴더 안에 1.png 부터 순서대로",),
        ("  사이즈:     size.png (고정 파일명)",),
        ("  영상:       video.mp4 (고정 파일명, mp4 권장)",),
        ("",),
        ("● products 시트 사용 규칙 (반드시 지켜주세요)",),
        ("",),
        ("  · no 컬럼은 시스템 ID — 절대 변경하지 마세요 (catalog.json 과 매칭되는 키)",),
        ("  · 중간에 행 추가·삭제 금지 — 새 상품은 가장 마지막 행 다음에 추가 (no = 334 부터)",),
        ("  · folder_name 은 영문 소문자 + 언더스코어만 (한글·공백·하이픈 X)",),
        ("  · main_image / detail_path / detail_url 은 folder_name 으로부터 자동 결정 — 수동 편집 불필요",),
        ("  · 시트 자체는 협업·확인용 문서 — 사이트 데이터는 catalog.json 이 정본",),
        ("",),
        ("● 작업 → 사이트 반영 흐름",),
        ("",),
        ("  1. 위 폴더에 파일 추가/수정",),
        ("  2. npm run dev (또는 npm run build) — sync-images 가 자동 실행되어",),
        ("     catalog.json 의 image/gallery/details/sizeImage/video 필드가 자동 갱신됩니다",),
        ("  3. git add . / git commit / git push",),
        ("  4. 1~2분 후 detail_url 에서 확인",),
        ("",),
        ("● 자주 묻는 질문",),
        ("",),
        ("  Q. 폴더 안에 파일을 넣었는데 사이트에 안 보입니다.",),
        ("  A. npm run sync-images 실행 후 catalog.json 변경분도 같이 commit 했는지 확인.",),
        ("     1~2분 정도 빌드/배포 시간 필요. 그래도 안 보이면 파일명 확인 (오타·확장자).",),
        ("",),
        ("  Q. 갤러리 사진 갯수에 제한이 있나요?",),
        ("  A. 제한 없음. 2.png 부터 빈 번호 없이 연속으로 추가.",),
        ("     중간에 빈 번호 있으면 그 앞까지만 인식됨 (예: 2,3,5 → 2,3 만 사용).",),
        ("",),
        ("  Q. 상세 이미지 1번 위에 헤더, 마지막에 푸터 같은 게 있어야 하나요?",),
        ("  A. 자유. 사이트는 1.png 부터 순서대로만 보여줍니다. 디자인은 자유.",),
        ("",),
        ("  Q. 모든 제품에 영상이나 사이즈 차트가 필요한가요?",),
        ("  A. 아니요. 있으면 표시, 없으면 그 영역 자체가 안 보임. 옵션입니다.",),
        ("",),
        ("  Q. npm run sync-images 가 무슨 일을 하나요?",),
        ("  A. catalog/ 폴더를 스캔해서 각 제품의 image/gallery/details/sizeImage/video",),
        ("     필드를 catalog.json 에 자동으로 채워줍니다. npm run dev/build 직전 자동 실행.",),
    ]

    for i, row in enumerate(rows, start=2):
        text = row[0]
        cell = ws.cell(row=i, column=1, value=text)
        if text.startswith("●"):
            cell.font = section_font
            cell.fill = section_fill
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)
        elif text.startswith("  D:") or text.startswith("  ├") or text.startswith("  │") or text.startswith("  └") or text.startswith("      "):
            cell.font = code_font
        else:
            cell.font = body_font
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        ws.row_dimensions[i].height = 18

    ws.column_dimensions["A"].width = 120


def build_cats_sheet(ws):
    """원본 Excel 에서 12개 고양이 상품 추출 → cats 시트"""
    headers = ["original_no", "brand", "product_name", "folder_name", "note"]
    write_header(ws, headers, fill_color="9CA3AF")  # 회색 — 비활성
    border = get_borders()

    cat_items = []
    try:
        wb_src = openpyxl.load_workbook(ORIGINAL_EXCEL, data_only=True)
        ws_src = wb_src["전체_세분류"]
        for row in ws_src.iter_rows(min_row=3, values_only=True):
            no = row[0]
            if not isinstance(no, int) or no not in CAT_NUMBERS:
                continue
            cat_items.append({
                "no": no,
                "brand": row[1] or "",
                "name": row[11] or "",
            })
    except Exception as ex:
        print(f"⚠ 원본 Excel 읽기 실패: {ex}")
        return 0

    # 고양이 폴더명도 비워두지 말고 영문 약어 부여 (작업자가 채울 자리)
    # 추후 활성화하려면 이 시트에서 데이터 갖고 와서 강아지 시트로 옮기면 됨
    for c in sorted(cat_items, key=lambda x: x["no"]):
        row_idx = ws.max_row + 1
        for col_idx, value in enumerate(
            [c["no"], c["brand"], c["name"], "", "고양이 상품 - 보류 (향후 활성화 시 products 시트로 이동)"],
            start=1,
        ):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.font = Font(color="6B7280")  # 회색 — 비활성 표시
            cell.alignment = Alignment(vertical="center")

    widths = {1: 12, 2: 16, 3: 60, 4: 32, 5: 30}
    for col_idx, w in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24

    return len(cat_items)


def main():
    catalog = json.loads(CATALOG_PATH.read_text(encoding="utf-8"))
    folder_list = json.loads(FOLDER_LIST_PATH.read_text(encoding="utf-8"))

    excel_by_norm = {}
    for e in folder_list:
        k = norm(e["name"])
        if k not in excel_by_norm:
            excel_by_norm[k] = e["folder"]

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # 시트 1: products (333개 강아지 상품)
    ws_products = wb.active
    ws_products.title = "products"
    matched = build_products_sheet(ws_products, catalog, excel_by_norm)

    # 시트 2: cats (12개 고양이 상품, 별도 분류)
    ws_cats = wb.create_sheet(title="cats")
    cat_count = build_cats_sheet(ws_cats)

    # 시트 3: guide (명명 규칙 + 폴더 구조 안내)
    ws_guide = wb.create_sheet(title="guide")
    build_guide_sheet(ws_guide)

    wb.save(OUT_PATH)

    # 결과 출력 (cp949 호환 — 이모지 X)
    print(f"[OK] product_list.xlsx 생성 완료")
    print(f"  - 위치: {OUT_PATH}")
    print(f"  - products 시트: {len(catalog)} 행 (no=1~{len(catalog)})")
    print(f"  - products 시트 folder_name 매칭됨: {matched}")
    print(f"  - cats sheet: {cat_count} rows")
    print(f"  - guide sheet: 폴더 구조 + 명명 규칙 안내")


if __name__ == "__main__":
    main()
